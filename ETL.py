import openpyxl
from openpyxl.utils import get_column_letter
import re
import json, jsonpickle
import datetime
import csv
import subprocess


def write_to_clipboard(output):
    process = subprocess.Popen(
        'pbcopy', env={'LANG': 'en_US.UTF-8'}, stdin=subprocess.PIPE)
    process.communicate(output.encode('utf-8'))

class Sheet:
    def __init__(self, name, configs = [], type='table'):
        self.name = name
        self.configs = configs
        self.type = type

class Config:
    def __init__(self, table='', start=0, end=None, direct=None, columns=[], extra=[], insert=[]):
        self.table = 'CAT_' + table.replace(' ','_') + '_Model__c'
        self.label = 'CAT ' + table + ' Model'
        self.start = start
        self.end = end
        self.relationship_name = table.replace(' ','_') + 's'
        self.direct = direct
        self.fields = []
        self.extra = extra
        self.columns = columns
        self.insert = insert
        self.sheet_name = None


    def is_direct(self):
        return self.direct != None

    def is_title(self, n):
        if self.is_direct(): return True
        return n == self.start

    def in_range(self, n):
        if n < self.start + 1: return False
        if self.end is None: return True
        return n <= self.end

    def get_field(self, sheet, row):
        if not self.is_direct():
            for i, field in enumerate(row):
                if i+1 in self.insert:
                    replacement = self.insert[i+1]
                    replacement.merge(row[i])
                    yield replacement
                else:
                    yield field
        else:
            for field in self.direct:
                ref = sheet[field.coordinate]
                field.data_type = ref.data_type
                yield field

    def skip_field(self, field):
        if not hasattr(field, 'column'): return False
        if not bool(self.columns): return False
        return not field.column in self.columns

    def should_skip(self, record):
        if self.table == 'CAT_Weighted_for_HCM_Model__c':
            if record['A_Customer_ID__c'] == '0':
                return True
        return False



class Field:
    def __init__(self, field_name=None, coordinate=None, value=None, distribution=None ):
        self.value = field_name
        self.coordinate = coordinate
        self.data_type = 's'
        self._value = value
        #self.cell = None
        self.number_format = 'General'
        self.column = None
        if distribution:
            #self.cell = distribution.cell
            self.column = distribution.cell.column
            self.value = distribution.cell.value
            self.coordinate = distribution.cell.coordinate
            self.data_type = distribution.get_type()
            self.col = re.findall('[A-Z]+', distribution.cell.coordinate)[0] if distribution.cell.coordinate else ''
            self.name = self.__get_name(self.col, distribution.cell.value,'__c')
            self._value = distribution.cell._value
            self.number_format = distribution.cell.number_format
            self.desc = 'Column: ' + self.col
            if distribution.cell.value:
                self.label = distribution.cell.value[0:40]


    def __hash__(self):
        if not self.coordinate:
            return hash(self.name)
        return hash(self.coordinate)

    def get_coordinate(self, row_number):
        if not self.coordinate:
            return
        if row_number is None:
            return self.coordinate
        return self.col + str(row_number)

    def __get_name(self, prefix, fieldname, postfix):
        if prefix:
            prefix = prefix + '_'

        name = re.sub('[^0-9a-zA-Z]+', '_', str(fieldname)).replace('__','_')
        if name == '':
            name = 'X'
        return prefix + name[0:33].strip('_') + postfix

    def merge(self, cell):
        if not cell: return
        #self.cell = cell
        self.coordinate = cell.coordinate
        self.col = re.findall('[A-Z]+', cell.coordinate)[0] + '_' if cell.coordinate else ''

special = 'System Costs' #this table is generated on the fly based upon selected hospital/group


class Distribution:
    def __init__(self, cell):
        self.cell = cell
        self.type = {}
        self.TYPE_STRING = 's'
        self.TYPE_FORMULA = 'f'
        self.TYPE_NUMERIC = 'n'
        self.TYPE_BOOL = 'b'
        self.TYPE_NULL = 'n'

    def get_type(self):
        c = 0
        rs = 's'
        for t in self.type:
            if self.type[t] >= c:
                c = self.type[t]
                rs = t
        return rs




class ETL:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self.configs = [
            Sheet('Toggles', [Config('Toggles',0, direct=[
                Field('Type O-neg Valuation Type', 'C3'),
                Field('Type O-neg Valuation Cost', 'D3'),
                Field('ABC Product Cost','C4'),
                Field('Indirect Cost to include','C5'),
                Field('Direct Product Cost Increase','C6'),
                Field('Pediatrics Handling Fee','C7'),
                Field('Costing:','C8'),
                Field('Marginal Cost Constrained By','C9'),
                Field('Percentage of O-neg Collected','C10'),
                Field('Avg Workload Time (min)','G3'),
                Field('Avg Non-Productive Time','G4'),
                Field('Avg Salary','G5'),
                Field('RBC Outdates','J4'),
                Field('SDP Outdates','J5'),
                Field('RBC Failed','K4'),
                Field('SDP Failed','K5'),
                Field('RBC Total','L4'),
                Field('SDP Total','L5'),
                Field('Handling Costs','L6'),
                Field('RBC Direct','J8'),
                Field('SDP Direct','J9'),
                Field('RBC Indirect','K8'),
                Field('SDP Indirect','K9'),
                Field('RBC Total','L8'),
                Field('SDP Total','L9'),
                Field('RBC Direct','O4'),
                Field('SDP Direct','O5'),
                Field('RDP Direct','O6'),
                Field('PSP Direct','O7'),
                Field('Plasma Direct','O8'),
                Field('Cryo Direct','O9'),
                Field('Cryo pools Direct','O10'),
                Field('RBC Direct','P4'),
                Field('SDP Direct','P5'),
                Field('RDP Direct','P6'),
                Field('PSP Direct','P7'),
                Field('Plasma Direct','P8'),
                Field('Cryo Direct','P9'),
                Field('Cryo pools Direct','P10'),
                Field('RBC Total','Q4'),
                Field('SDP Total','Q5'),
                Field('RDP Total','Q6'),
                Field('PSP Total','Q7'),
                Field('Plasma Total','Q8'),
                Field('Cryo Total','Q9'),
                Field('Cryo pools Total','Q10'),
                Field('Direct Cost Frac','T3'),
                Field('ASP Frac','T4'),
                Field('Margin Frac','T5'),
                Field('Percentage of Frac Producing RBC units','T6')
            ])], type='map'),
            Sheet('SDP RBC Weighted for HCM', [
                Config('Weighted for HCM',4, extra=[Field('Type', None, value='RBC')], columns=[1,2]),
                Config('Weighted for HCM',4, extra=[Field('Type', None, value='SDP')], columns=[4,5])
            ]),
            Sheet('Weighted Dist', [Config('Weighted_Dist',3)]),
            Sheet('Logistics Tiers', [Config('Logistics_Tiers',1, columns=[2,3,4,5,6,7,8,9,10,11,12,13])]),
            Sheet('IRL', [Config('IRL',3, columns=[2,3,4,5,6,7,8,9,10,11,12,13,14,15],
                                 insert={2:Field('IRL Facility','B4'),
                                         3:Field('City','C4'),
                                         4:Field('Region','D4'),
                                         5:Field('Region','E4'),
                                         13:Field('IRL','M4'),
                                         14:Field('Molecular','N4'),
                                         })]),
            Sheet('Distribution', [Config('Distribution',1,columns=[2,3,4,5,6,7,8,8,9,10,11,12,13,14,15,16,17])]),
            Sheet('Distribution - Logistics', [Config('Distribution_Logistics',1)]),
            Sheet('Other Rev Margin', [Config('Other Rev Margin',0,direct=[
                Field('Other Rev','B3'),
                Field('Total RDP Revenue','B5'),
                Field('Percent of Other Rev RDP Revenue','B6'),
                Field('System Margin RDP Revenue','B7'),
                Field('Sumprod RDP Revenue','B8'),
                Field('RDP ASP RDP Revenue','B10'),
                Field('RDP COST RDP Revenue','B11'),
                Field('Total PACS Revenue','C5'),
                Field('Percent of Other Rev PACS Revenue','C6'),
                Field('System Margin PACS Revenue','C7'),
                Field('Total Stem Revenue','D5'),
                Field('Percent of Other Rev Stem Revenue','D6'),
                Field('System Margin Stem Revenue','D7'),
                Field('Total Other Products Revenue','E5'),
                Field('Percent of Other Rev Other Products Revenue','E6'),
                Field('System Margin Other Products Revenue','E7'),
                Field('Total Non-Product Revenue','F5'),
                Field('Percent of Other Rev Non-Product Revenue','F6'),
                Field('System Margin Non-Product Revenue','F7')
            ])], type='map'),

            Sheet('Revenue', [Config('Revenue',0)]),
            Sheet('Map P2', [Config('Map P2',0,None)]),
            Sheet('Weights', [Config('Weights',0)])]
    meta = '''<?xml version="1.0" encoding="UTF-8"?>
<CustomObject xmlns="http://soap.sforce.com/2006/04/metadata">
    <label>%s</label>
    <nameField>
        <label>Row Number</label>
        <trackHistory>false</trackHistory>
        <type>Text</type>
    </nameField>
    <pluralLabel>%s</pluralLabel>
    <sharingModel>ControlledByParent</sharingModel>
    <visibility>Public</visibility>
    <deploymentStatus>Deployed</deploymentStatus>
    <fields>
        <fullName>CAT_Model__c</fullName>
        <externalId>false</externalId>
        <label>Cat Model</label>
        <referenceTo>CAT_Model__c</referenceTo>
        <relationshipLabel>CAT Models</relationshipLabel>
        <relationshipName>%s</relationshipName>
        <relationshipOrder>0</relationshipOrder>
        <reparentableMasterDetail>true</reparentableMasterDetail>
        <trackHistory>false</trackHistory>
        <trackTrending>false</trackTrending>
        <type>MasterDetail</type>
        <writeRequiresMasterRead>false</writeRequiresMasterRead>
    </fields>
    <fields>
        <fullName>Row__c</fullName>
        <externalId>false</externalId>
        <label>Row</label>
        <precision>18</precision>
        <required>true</required>
        <scale>0</scale>
        <trackTrending>false</trackTrending>
        <type>Number</type>
        <unique>false</unique>
    </fields>'''



    def text_field(self, name, desc, label, length=50, externalId = False ):
        rs = '''
    <fields>
        <fullName>%s</fullName>
        <description>%s</description>
        <externalId>%s</externalId>
        <label>%s</label>
        <length>%s</length>
        <required>false</required>
        <trackFeedHistory>false</trackFeedHistory>
        <trackHistory>false</trackHistory>
        <type>Text</type>
        <unique>false</unique>
    </fields>''' % (name, desc, str(externalId).lower(), label, length)
        return rs

    def number_field(self, config, sheet, col, externalId = False ):
        field = Field(distribution=col)
        rs = '''
    <fields>
        <fullName>%s</fullName>
        <description>%s</description>
        <externalId>%s</externalId>
        <label>%s</label>
        <precision>16</precision>
        <required>false</required>
        <scale>2</scale>
        <trackFeedHistory>false</trackFeedHistory>
        <trackHistory>false</trackHistory>
        <type>Number</type>
        <unique>false</unique>
    </fields>''' % (field.name, field.desc, str(externalId).lower(), field.label)
        return rs

    def checkbox_field(self, config, sheet, col ):
        field = Field(distribution=col)
        rs = '''
        <fields>
            <fullName>%s</fullName>
            <defaultValue>false</defaultValue>
            <description>%s</description>
            <externalId>false</externalId>
            <label>%s</label>
            <trackFeedHistory>false</trackFeedHistory>
            <trackHistory>true</trackHistory>
            <type>Checkbox</type>
        </fields>''' % (field.name, field.desc, field.label)
        return rs

    def field_map(self):
        for config in self.configs:
            sheet = self.wb[config.name]
            for idx, cfg in enumerate(config.configs):
                meta = ETL.meta% (cfg.label, cfg.table, cfg.relationship_name)
                for col in self.get_distributions(config, sheet):
                    type = col.get_type()
                    if not col.cell.value:
                        continue
                    field = Field(distribution=col)
                    print("'%s' => '%s',"%(field.coordinate, field.name))


    def generate_table_name(self, sheet_name):
        name = re.sub('[^A-Za-z0-9]','', sheet_name)
        return name

    def clean(self, value):
        if value is None:
            return ''
        return value

    def generate_object_meta(self):

        for sheet_name in self.wb.sheetnames:
            if sheet_name in ('System Costs','Home','Inputs','Sheet2','RBC Returns Old','SDP Returns Old'): continue
            sheet = self.wb[sheet_name]
            table = 'CAT_Model_' + self.generate_table_name(sheet_name) + '__c'
            out = open('output/cat/' + table  + '.object','w')
            data_out = csv.writer(open('output/data/' + table  + '.csv','w'))
            meta = ETL.meta% (sheet_name, sheet_name, self.generate_table_name(sheet_name)  )

            for idx, row in enumerate(sheet.iter_rows()):
                if idx == 0:
                    columns = ['Name','CAT_Model__c','Row__c']
                    for i, column in enumerate(row):
                        col_name = get_column_letter(i + 1)
                        meta += self.text_field(col_name+'__c', '',col_name )
                        columns.append(col_name+'__c')
                    data_out.writerow(columns)

                else:
                    data_out.writerow([str(idx + 1), '', str(idx + 1)] + [self.clean(v.value) for v in row])


            out.write(meta)
            out.write('\n</CustomObject>')
            print('output: ' + table)

    def create_model_record(self, sf, label, type, is_active):
        rs = sf.CAT_Model__c.create({'Name' : label[0:80], 'Active__c' : is_active})
        if not rs['success']: raise Exception('couldn\'t create record')
        return rs['id']

    def execute(self, sf, table, data, cleanup = False):
        print('executing', table, datetime.datetime.now())
        rs = sf.bulk.__getattr__(table).insert(data)
        for r in rs:
            if not r['success']:
                print(r)
        print('execution complete', datetime.datetime.now())
        return True

    def transfer(self, sf):
        print('start', datetime.datetime.now())
        name = self.wb['Inputs']['B8'].value
        model_id = self.create_model_record(sf, name, '',True)

        for sheet_name in self.wb.sheetnames:
            if sheet_name in ('System Costs','Home','Inputs','Sheet2','RBC Returns Old','SDP Returns Old'): continue
            sheet = self.wb[sheet_name]
            table = 'CAT_Model_' + self.generate_table_name(sheet_name) + '__c'
            data = []
            for i, row in enumerate(sheet.iter_rows()):
                copy = {}
                copy['CAT_Model__c'] = model_id
                copy['Row__c'] = str(i+1)
                copy['Name'] = copy['Row__c']
                for n, cell in enumerate(row):
                    if cell.value:
                        value = str(cell.value)[0:50]
                    else:
                        value = None

                    copy[get_column_letter(n+1) + '__c']= value
                data.append(copy)
            self.execute(sf, table, data, True)

    def data_map(self):
        name = self.wb['Inputs']['B8'].value
        result = ''

        for sheet_name in self.wb.sheetnames:
            if sheet_name in ('System Costs','Home','Inputs','Sheet2','RBC Returns Old','SDP Returns Old'): continue
            print(sheet_name)
            sheet = self.wb[sheet_name]
            table = 'Salesforce Table: CAT_Model_' + self.generate_table_name(sheet_name) + '__c'
            result += table + '\n'
            result += 'Excel Tab: ' + sheet_name + '\n'
            data = []
            result += 'Fields: '
            for i, row in enumerate(sheet.iter_rows()):
                result += 'Name '
                result += 'Row__c '

                for n, cell in enumerate(row):
                    result +=  get_column_letter(n+1) + '__c' + ' '
                result += '\n\n'

                break
        print(result)
        write_to_clipboard(result)



    def transform(self, sheet, cfg, model_id, idx, data):
        rs = {}
        rs['Cat_Model__c'] = model_id
        rs['Row__c'] = idx
        if cfg.type == 'table':
            for field in cfg.fields:
                if field.column:
                    for cell in data:
                        if field.column == cell.column:
                            rs[field.name] =  self.clean_value(cfg, cell.value)
                            break
                else:
                    rs[field.name] = self.clean_value(cfg, field._value)
        elif cfg.type == 'map':
            for field in cfg.fields:
                coord = field.get_coordinate(None)
                if coord:
                    cell = sheet[coord]
                    rs[field.name] =  self.clean_value(cfg, cell.value)
                else:
                    rs[field.name] = self.clean_value(cfg, field._value)
        return rs

    def clean_value(self, cfg, value):
        if value is None: return value
        if isinstance(value, str):
            if value.isnumeric(): round(float(value),2)
        return value



    def get_distributions(self, config, sheet, get_distribution = True):
        for idx, cfg in enumerate(config.configs):
            anal = {}
            if cfg.is_direct():
                for field in cfg.direct:
                    anal[field.coordinate] = Distribution(field)
                    anal[field.coordinate].type[field.data_type] = 1
            else:
                for i, row in enumerate(sheet.iter_rows()):
                    if not cfg.is_title(i) and not cfg.in_range(i): continue
                    for field in cfg.get_field(sheet, row):
                        if cfg.skip_field(field): continue
                        if cfg.is_title(i):
                            if isinstance(field, openpyxl.cell.read_only.EmptyCell): continue
                            key = re.findall('[A-Z]+', field.coordinate)[0]
                            anal[key] = Distribution(field)
                        elif cfg.in_range(i):
                            if not get_distribution: break
                            if field.value is None: continue
                            key = re.findall('[A-Z]+', field.coordinate)[0]
                            if key in anal:
                                if field.data_type in anal[key].type:
                                    anal[key].type[field.data_type] += 1
                                else:
                                    anal[key].type[field.data_type] = 1

                for field in cfg.extra:
                    anal[field.value] = Distribution(field)

            return anal.values()


    def generate_config(self, output_path, type='JSON'):
        output = []
        for config in self.configs:
            sheet = self.wb[config.name]
            for c in config.configs:
                c.sheet_name = config.name
                c.type = config.type
                for dist in self.get_distributions(config, sheet, True):
                    f= Field(distribution=dist)
                    c.fields.append(f)
                output.append(c)
        open(output_path, 'w').write((json.dumps(json.loads(jsonpickle.encode(output)), indent=4)))
